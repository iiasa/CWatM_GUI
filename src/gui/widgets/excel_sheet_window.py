"""
Excel sheet viewer/editor for the CWatM GUI (Excel menu).

``ExcelSheetWindow`` opens one worksheet of an ``.xlsx`` workbook (e.g. the
``Crops``/``Reservoirs`` sheets of the settings ``Excel_settings_file``) in an
**editable table that reproduces the sheet's cell colours** (fill + font) with
openpyxl. Bottom buttons: **Reload / Save / Save As** (and, for Reservoirs, a
**Release** button opening the ``Reservoirs_downstream`` companion sheet).

Rendering uses a **lazy** ``QTableView`` + ``ExcelSheetModel``
(``QAbstractTableModel``): the view only asks the model for the cells that are
actually visible, so even a 300k+-cell sheet opens instantly and scrolls smoothly
(no per-cell widgets are built up-front). Saving writes the edited values back
through openpyxl so every other sheet and all the styling is preserved.
"""

import os

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableView,
    QFileDialog, QMessageBox, QAbstractItemView,
)
from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex
from PySide6.QtGui import QColor, QBrush, QFont, QIcon

from src.gui.utils import theme
from src.gui.utils.window_geometry import GeometryMemoryMixin
from src.gui.utils.gui_log import get_logger

log = get_logger("excel_sheet_window")


def _argb_to_qcolor(argb):
    """openpyxl ARGB string ('FF00A87C' or '00A87C') -> QColor, or None if it is not
    a plain rgb colour (theme/indexed colours are left uncoloured)."""
    if not isinstance(argb, str):
        return None
    s = argb.strip()
    if len(s) == 8:
        a, rgb = int(s[0:2], 16), s[2:]
    elif len(s) == 6:
        a, rgb = 255, s
    else:
        return None
    try:
        r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return None
    if a == 0:            # fully transparent fill == "no fill"
        return None
    return QColor(r, g, b)


def _parse_cell(text):
    """Turn edited cell text back into a number when it looks like one (blank -> None),
    so the saved xlsx keeps numeric types."""
    t = text.strip()
    if t == "":
        return None
    try:
        return int(t)
    except ValueError:
        pass
    try:
        return float(t)
    except ValueError:
        pass
    return text


class ExcelSheetModel(QAbstractTableModel):
    """Lazy model over an openpyxl worksheet: values, per-cell fill/font colours and
    editing are served on demand (only for the cells the view actually shows)."""

    def __init__(self, ws, parent=None):
        super().__init__(parent)
        self._ws = ws
        self._nrows = max(ws.max_row, 1)
        self._ncols = max(ws.max_column, 1)
        self._edits = {}          # (row0, col0) -> new text (differs from the cell)
        self._style_cache = {}    # (row0, col0) -> (bg QBrush|None, fg QBrush|None, QFont|None)

    # ---- shape
    def rowCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else self._nrows

    def columnCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else self._ncols

    # ---- helpers
    def _cell(self, r, c):
        return self._ws.cell(row=r + 1, column=c + 1)

    def _orig_text(self, r, c):
        v = self._cell(r, c).value
        return "" if v is None else str(v)

    def _display_text(self, r, c):
        """Text shown in the (non-edited) cell: floats are limited to **max 5
        decimals** (trailing zeros stripped); everything else is shown verbatim."""
        v = self._cell(r, c).value
        if isinstance(v, float):
            s = f"{v:.5f}".rstrip("0").rstrip(".")
            return "0" if s in ("", "-", "-0") else s
        return "" if v is None else str(v)

    def _style(self, r, c):
        key = (r, c)
        cached = self._style_cache.get(key)
        if cached is not None:
            return cached
        bg = fg = font = None
        try:
            cell = self._cell(r, c)
            fill = cell.fill
            if fill is not None and fill.patternType == "solid":
                col = _argb_to_qcolor(getattr(fill.fgColor, "rgb", None))
                if col is not None:
                    bg = QBrush(col)
            f = cell.font
            if f is not None:
                fcol = _argb_to_qcolor(getattr(getattr(f, "color", None), "rgb", None))
                if fcol is not None:
                    fg = QBrush(fcol)
                if f.bold or f.italic:
                    font = QFont()
                    font.setBold(bool(f.bold))
                    font.setItalic(bool(f.italic))
        except Exception:
            log.debug("excel cell style failed", exc_info=True)
        val = (bg, fg, font)
        self._style_cache[key] = val
        return val

    # ---- data
    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        r, c = index.row(), index.column()
        if role == Qt.DisplayRole:
            if (r, c) in self._edits:
                return self._edits[(r, c)]
            return self._display_text(r, c)          # max 5 decimals for floats
        if role == Qt.EditRole:
            if (r, c) in self._edits:
                return self._edits[(r, c)]
            return self._orig_text(r, c)             # full precision while editing
        if role == Qt.BackgroundRole:
            return self._style(r, c)[0]
        if role == Qt.ForegroundRole:
            return self._style(r, c)[1]
        if role == Qt.FontRole:
            return self._style(r, c)[2]
        return None

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable

    def setData(self, index, value, role=Qt.EditRole):
        if role != Qt.EditRole or not index.isValid():
            return False
        r, c = index.row(), index.column()
        txt = "" if value is None else str(value)
        if txt == self._orig_text(r, c):
            self._edits.pop((r, c), None)        # reverted -> no longer an edit
        else:
            self._edits[(r, c)] = txt
        self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
        return True

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        from openpyxl.utils import get_column_letter
        return get_column_letter(section + 1) if orientation == Qt.Horizontal \
            else str(section + 1)

    # ---- edits / save
    def has_edits(self):
        return bool(self._edits)

    def flush_to_wb(self):
        """Write the changed cells into the worksheet (numbers re-parsed); unchanged
        cells - including \\xa0 spacers - are left untouched so their formatting
        survives. Merged non-anchor cells are skipped."""
        for (r, c), txt in list(self._edits.items()):
            try:
                self._ws.cell(row=r + 1, column=c + 1).value = _parse_cell(txt)
            except Exception:
                log.debug("excel write cell (%d,%d) failed", r, c, exc_info=True)
        self._edits.clear()


class _FrozenRowTableView(QTableView):
    """QTableView with the sheet's **first data row (row 0) frozen** at the top.

    A small second ``QTableView`` (``self._frozen``) shares the same model and is
    stacked over the main viewport, sized to exactly one row and pinned just below
    the column-letter header, so row 0 stays visible while the main view scrolls
    down. Horizontal scrolling and column widths are kept in sync; the frozen strip
    is read-only (row 0 is a header row in these sheets)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._frozen = QTableView(self)
        f = self._frozen
        f.setFocusPolicy(Qt.NoFocus)
        f.horizontalHeader().hide()
        f.setEditTriggers(QAbstractItemView.NoEditTriggers)
        f.setSelectionMode(QAbstractItemView.NoSelection)
        f.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        f.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        f.setStyleSheet("QTableView { border: none; }")
        f.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        f.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)

        self.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.viewport().stackUnder(f)

        self.horizontalHeader().sectionResized.connect(self._on_section_resized)
        self.horizontalScrollBar().valueChanged.connect(
            f.horizontalScrollBar().setValue)
        # the main vertical header width can grow while scrolling (wider row
        # numbers) -> keep the frozen strip aligned.
        self.verticalScrollBar().valueChanged.connect(self._update_frozen_geometry)

    def setModel(self, model):
        super().setModel(model)
        self._frozen.setModel(model)
        self._frozen.setShowGrid(self.showGrid())
        self._frozen.show()
        self._update_frozen_geometry()

    def sync_frozen_columns(self):
        """Copy the main view's column widths onto the frozen strip (call after the
        main view auto-sizes / sets a default width)."""
        if self.model() is None:
            return
        self._frozen.horizontalHeader().setDefaultSectionSize(
            self.horizontalHeader().defaultSectionSize())
        for c in range(self.model().columnCount()):
            self._frozen.setColumnWidth(c, self.columnWidth(c))
        self._update_frozen_geometry()

    def _on_section_resized(self, logical, old, new):
        self._frozen.setColumnWidth(logical, new)
        self._update_frozen_geometry()

    def _update_frozen_geometry(self, *_):
        model = self.model()
        if model is None:
            return
        vh_w = self.verticalHeader().width()
        self._frozen.verticalHeader().setFixedWidth(vh_w)
        row_h = self.rowHeight(0) if model.rowCount() > 0 \
            else self.verticalHeader().defaultSectionSize()
        self._frozen.setGeometry(
            self.frameWidth(),
            self.horizontalHeader().height() + self.frameWidth(),
            vh_w + self.viewport().width(),
            row_h)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_frozen_geometry()


class ExcelSheetWindow(GeometryMemoryMixin, QDialog):
    """Editable, colour-preserving view of one worksheet (lazy QTableView)."""

    def __init__(self, path, sheet_name, parent=None, release_sheet=None):
        super().__init__(parent)
        self.path = path
        self.sheet_name = sheet_name
        # Optional companion sheet reachable via a "Release" button (Reservoirs ->
        # Reservoirs_downstream); greyed out when that sheet is absent.
        self._release_sheet = release_sheet
        self._wb = None
        self._ws = None
        self.model = None

        self.setWindowTitle(f"{sheet_name} — {os.path.basename(path)}")
        self.setModal(True)
        self.setWindowFlags(Qt.Dialog | Qt.WindowMinMaxButtonsHint
                            | Qt.WindowCloseButtonHint)
        if not self._init_geometry_memory(f"excel_{sheet_name.lower()}"):
            self.resize(900, 600)
        try:
            icon = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(
                os.path.dirname(__file__)))), 'assets', 'cwatm.ico')
            if os.path.exists(icon):
                self.setWindowIcon(QIcon(icon))
        except Exception:
            pass

        self._build_ui()
        self._load()

    # ------------------------------------------------------------------- UI
    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(8)

        self.header_label = QLabel(f"Sheet: {self.sheet_name}")
        self.header_label.setStyleSheet(
            "font-family:'Segoe UI',sans-serif; font-size:14px; font-weight:600; "
            f"color:{theme.c('text')}; padding:2px;")
        lay.addWidget(self.header_label)

        self.table = _FrozenRowTableView()
        self.table.setEditTriggers(
            QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed)
        # Lazy: only visible cells are queried. Sample few rows when auto-sizing so a
        # huge sheet does not scan every row.
        self.table.horizontalHeader().setResizeContentsPrecision(40)
        lay.addWidget(self.table, 1)

        self.info_label = QLabel("")
        self.info_label.setStyleSheet(
            f"font-size:12px; color:{theme.c('text_muted')};")
        lay.addWidget(self.info_label)

        _btn = """
            QPushButton { font-family:'Segoe UI',sans-serif; font-size:12px;
                font-weight:500; color:white; border:none; border-radius:6px;
                padding:6px 16px; min-height:26px;
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #5dade2, stop:1 #3498db); }
            QPushButton:hover { background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                stop:0 #85c1e9, stop:1 #5dade2); }
            QPushButton:disabled { background:#d3d3d3; color:#a9a9a9; }
        """
        _gray = _btn.replace("#5dade2", "#808080").replace("#3498db", "#606060") \
                    .replace("#85c1e9", "#a0a0a0")

        row = QHBoxLayout()
        row.setSpacing(10)
        # "Load": open ANOTHER workbook in this same editor, showing the same sheet
        # (Crops / Reservoirs) - so a different xlsx than the settings
        # Excel_settings_file can be inspected without changing the settings.
        self.load_button = QPushButton("Load")
        self.load_button.setStyleSheet(_btn)
        self.load_button.setToolTip("load an Excel file")
        self.load_button.clicked.connect(self._load_other)
        row.addWidget(self.load_button)

        self.reload_button = QPushButton("Reload")
        self.reload_button.setStyleSheet(_btn)
        self.reload_button.setToolTip("Discard edits and reload the sheet from disk")
        self.reload_button.clicked.connect(self._reload)
        row.addWidget(self.reload_button)

        self.save_button = QPushButton("Save")
        self.save_button.setStyleSheet(_btn)
        self.save_button.setToolTip("Write the edits back to the Excel file")
        self.save_button.clicked.connect(self._save)
        row.addWidget(self.save_button)

        self.save_as_button = QPushButton("Save As")
        self.save_as_button.setStyleSheet(_btn)
        self.save_as_button.clicked.connect(self._save_as)
        row.addWidget(self.save_as_button)

        # "Release": open a companion sheet in its own editor (greyed out if absent).
        self.release_button = None
        if self._release_sheet:
            row.addSpacing(28)
            self.release_button = QPushButton("Release")
            self.release_button.setStyleSheet(_btn)
            self.release_button.setToolTip(
                f"Open the '{self._release_sheet}' sheet in a new window")
            self.release_button.setEnabled(False)   # enabled in _load if it exists
            self.release_button.clicked.connect(self._open_release)
            row.addWidget(self.release_button)

        row.addStretch()
        self.close_button = QPushButton("Close")
        self.close_button.setStyleSheet(_gray)
        self.close_button.clicked.connect(self.close)
        row.addWidget(self.close_button)
        lay.addLayout(row)

    # ----------------------------------------------------------------- load
    def _load(self):
        """Load self.path into the table. Returns True on success - the Load button
        uses that to fall back to the previously shown workbook."""
        from openpyxl import load_workbook
        try:
            self._wb = load_workbook(self.path)   # keep styles (for colours + save)
        except Exception as e:
            QMessageBox.warning(self, self.sheet_name,
                                f"Could not open the Excel file:\n{e}")
            return False
        if self.release_button is not None:
            self.release_button.setEnabled(
                self._release_sheet in self._wb.sheetnames)
        if self.sheet_name not in self._wb.sheetnames:
            QMessageBox.warning(
                self, self.sheet_name,
                f"The workbook has no sheet '{self.sheet_name}'.\n"
                f"Sheets: {', '.join(self._wb.sheetnames)}")
            return False
        self._ws = self._wb[self.sheet_name]
        self.model = ExcelSheetModel(self._ws, self)
        self.table.setModel(self.model)
        nrows, ncols = self.model.rowCount(), self.model.columnCount()
        # Only auto-size columns for a modest sheet (still cheap thanks to the
        # resize precision); wide sheets keep a sensible default width.
        if ncols <= 60:
            self.table.resizeColumnsToContents()
        else:
            self.table.horizontalHeader().setDefaultSectionSize(90)
        self.table.sync_frozen_columns()      # keep the frozen first row aligned
        self.info_label.setText(f"{nrows} rows × {ncols} columns   |   {self.path}")
        return True

    # ----------------------------------------------------------------- save
    def _has_edits(self):
        return self.model is not None and self.model.has_edits()

    def _write(self, path):
        if self._wb is None or self.model is None:
            return False
        self.model.flush_to_wb()
        try:
            self._wb.save(path)
            return True
        except PermissionError:
            QMessageBox.warning(
                self, "Save",
                "Could not save - the file is open in Excel (or read-only).\n"
                "Close it there and try again.")
        except Exception as e:
            QMessageBox.warning(self, "Save", f"Could not save the file:\n{e}")
        return False

    def _save(self):
        if self._write(self.path):
            self.info_label.setText(f"Saved: {self.path}")

    def _save_as(self):
        start = os.path.splitext(self.path)[0] + "_edited.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel as", start, "Excel files (*.xlsx)")
        if not path:
            return
        if self._write(path):
            self.path = path
            self.setWindowTitle(f"{self.sheet_name} — {os.path.basename(path)}")
            self.info_label.setText(f"Saved: {path}")

    def _load_other(self):
        """Load button: pick another .xlsx and show ITS Crops / Reservoirs sheet in
        this window. The settings file is untouched - Save then writes to the newly
        loaded workbook, which is why unsaved edits are confirmed away first."""
        if self._has_edits():
            if QMessageBox.question(
                    self, "Load",
                    "Discard your edits and load another Excel file?") \
                    != QMessageBox.Yes:
                return
        start = os.path.dirname(self.path) if self.path else ""
        path, _ = QFileDialog.getOpenFileName(
            self, f"Load an Excel file ({self.sheet_name} sheet)", start,
            "Excel files (*.xlsx);;All files (*)")
        if not path:
            return
        previous = self.path
        self.path = path
        if self._load():
            self.setWindowTitle(f"{self.sheet_name} — {os.path.basename(path)}")
        elif previous and previous != path:
            # Unreadable file or no such sheet: go back to what was on screen, so
            # the window never ends up showing one workbook while self.path (what
            # Save writes to) points at another.
            self.path = previous
            self._load()
            self.setWindowTitle(
                f"{self.sheet_name} — {os.path.basename(previous)}")

    def _reload(self):
        if self._has_edits():
            if QMessageBox.question(
                    self, "Reload",
                    "Discard your edits and reload the sheet from disk?") \
                    != QMessageBox.Yes:
                return
        self._load()

    def _open_release(self):
        """Open the companion sheet (e.g. Reservoirs_downstream) in its own editor."""
        if not self._release_sheet:
            return
        try:
            win = ExcelSheetWindow(self.path, self._release_sheet, self)
            win.exec()
        except Exception as e:
            QMessageBox.warning(
                self, "Release",
                f"Could not open '{self._release_sheet}':\n{e}")
